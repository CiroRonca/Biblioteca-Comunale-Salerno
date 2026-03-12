#!/usr/bin/env python3
"""
Catalogatore Bibliografico
--------------------------
Legge un file PDF o TXT con record bibliografici e produce:
  - catalogo.xlsx  (tabella Excel formattata)
  - catalogo.csv   (tabella CSV)

Uso:
  python catalogo_pdf.py mio_file.pdf
  python catalogo_pdf.py mio_file.txt
"""

import re
import sys
import os
import pdfplumber
import pandas as pd
from pathlib import Path


# ── ESTRAZIONE TESTO ──────────────────────────────────────────────────────────

def estrai_testo_pdf(percorso: str) -> str:
    """Estrae tutto il testo da un PDF pagina per pagina."""
    testo = ""
    with pdfplumber.open(percorso) as pdf:
        totale = len(pdf.pages)
        for i, pagina in enumerate(pdf.pages, 1):
            t = pagina.extract_text()
            if t:
                testo += t + "\n"
            print(f"  Pagina {i}/{totale}...", end="\r")
    print()
    return testo


def estrai_testo_txt(percorso: str) -> str:
    """Legge un file di testo con rilevamento automatico dell'encoding."""
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            with open(percorso, encoding=enc) as f:
                return f.read()
        except UnicodeDecodeError:
            continue
    raise ValueError("Impossibile decodificare il file di testo.")


# ── PARSING DEI RECORD ────────────────────────────────────────────────────────

def parse_record(blocco: str) -> dict | None:
    """
    Parsifica un singolo blocco di testo bibliografico.

    Formato atteso:
      [M]-Titolo / Autore. - Città : Editore, anno. -
      NNN p. ; XX cm.
      CODICE

    Regole:
      - Titolo  → tutto prima dello "/"
      - Autore  → dopo "/" fino al primo " - "
      - Pagine  → numero prima di "p."
      - Codice  → tutto dopo "cm."
    """
    # Unifica su una sola riga e pulisci spazi multipli
    riga = " ".join(blocco.split())

    # Rimuovi prefisso tipo [M]- [P]- ecc.
    riga = re.sub(r"^\[.*?\]-?\s*", "", riga).strip()

    # Titolo: prima dello "/"
    idx_slash = riga.find("/")
    if idx_slash == -1:
        return None
    titolo = riga[:idx_slash].strip()

    # Autore: dopo "/" fino al primo " - "
    resto = riga[idx_slash + 1:].strip()
    idx_trattino = resto.find(" - ")
    autore = resto[:idx_trattino].strip() if idx_trattino != -1 else resto.strip()

    # Pagine: numero prima di "p."
    m_pagine = re.search(r"(\d+)\s*p\.", riga)
    pagine = m_pagine.group(1) if m_pagine else ""

    # Codice: tutto dopo "cm."
    idx_cm = riga.find("cm.")
    codice = riga[idx_cm + 3:].strip() if idx_cm != -1 else ""

    if not titolo and not autore:
        return None

    return {
        "Titolo":  titolo,
        "Autore":  autore,
        "Pagine":  f"{pagine} p." if pagine else "",
        "Codice":  codice,
    }


def parse_testo(testo: str) -> list[dict]:
    """Suddivide il testo in blocchi e parsifica ogni record."""
    # Divide sui marker [X] o su righe vuote doppie
    blocchi = re.split(r"\n(?=\[)|(?:\n\s*){2,}", testo)
    risultati = []
    for blocco in blocchi:
        blocco = blocco.strip()
        if not blocco:
            continue
        record = parse_record(blocco)
        if record:
            risultati.append(record)
    return risultati


# ── ESPORTAZIONE ──────────────────────────────────────────────────────────────

def esporta_excel(df: pd.DataFrame, percorso: str):
    """Crea un file Excel formattato con intestazioni colorate."""
    with pd.ExcelWriter(percorso, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Catalogo")

        ws = writer.sheets["Catalogo"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Intestazioni
        fill_header = PatternFill("solid", fgColor="1E1E30")
        font_header = Font(bold=True, color="B8963E", name="Courier New", size=10)
        border_bottom = Border(bottom=Side(style="medium", color="B8963E"))

        col_widths = {"Titolo": 55, "Autore": 30, "Pagine": 12, "Codice": 18}

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_bottom
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

        # Righe dati con colori alternati
        fill_pari   = PatternFill("solid", fgColor="0A0A14")
        fill_dispari = PatternFill("solid", fgColor="0D0D1C")
        font_titolo = Font(italic=True, color="E8E0D0", name="Georgia", size=10)
        font_autore = Font(color="7090C0", name="Courier New", size=9)
        font_pagine = Font(color="90B870", name="Courier New", size=9)
        font_codice = Font(color="B8963E", name="Courier New", size=9)
        font_stili  = [font_titolo, font_autore, font_pagine, font_codice]

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            fill = fill_pari if row_idx % 2 == 0 else fill_dispari
            for col_idx, cell in enumerate(row):
                cell.fill = fill
                cell.font = font_stili[col_idx]
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 0))

        # Blocca la riga di intestazione
        ws.freeze_panes = "A2"


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Uso: python catalogo_pdf.py <file.pdf o file.txt>")
        sys.exit(1)

    percorso_input = sys.argv[1]
    if not os.path.exists(percorso_input):
        print(f"Errore: file non trovato → {percorso_input}")
        sys.exit(1)

    estensione = Path(percorso_input).suffix.lower()
    nome_base  = Path(percorso_input).stem

    print(f"\n{'='*55}")
    print(f"  CATALOGATORE BIBLIOGRAFICO")
    print(f"{'='*55}")
    print(f"  File:  {percorso_input}")
    print(f"  Tipo:  {estensione.upper()[1:]}")
    print(f"{'='*55}\n")

    # 1. Estrai testo
    print("► Estrazione testo...")
    if estensione == ".pdf":
        testo = estrai_testo_pdf(percorso_input)
    elif estensione in (".txt", ".text"):
        testo = estrai_testo_txt(percorso_input)
    else:
        print("Errore: formato non supportato. Usa .pdf o .txt")
        sys.exit(1)

    # 2. Parsifica
    print("► Analisi record bibliografici...")
    records = parse_testo(testo)

    if not records:
        print("\n⚠  Nessun record riconosciuto nel file.")
        print("   Controlla che il formato rispetti lo schema:")
        print("   [M]-Titolo / Autore. - Città : Editore, anno. -")
        print("   NNN p. ; XX cm.")
        print("   CODICE")
        sys.exit(1)

    df = pd.DataFrame(records)

    # 3. Statistiche
    senza_pagine = df["Pagine"].eq("").sum()
    senza_codice = df["Codice"].eq("").sum()
    print(f"\n  Record trovati  : {len(df)}")
    print(f"  Senza pagine    : {senza_pagine}")
    print(f"  Senza codice    : {senza_codice}")

    # 4. Esporta
    out_dir = Path(percorso_input).parent

    percorso_xlsx = str(out_dir / f"{nome_base}_catalogo.xlsx")
    percorso_csv  = str(out_dir / f"{nome_base}_catalogo.csv")

    print("\n► Creazione Excel...")
    esporta_excel(df, percorso_xlsx)

    print("► Creazione CSV...")
    df.to_csv(percorso_csv, index=False, encoding="utf-8-sig")

    print(f"\n{'='*55}")
    print(f"  ✓  Excel → {percorso_xlsx}")
    print(f"  ✓  CSV   → {percorso_csv}")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()